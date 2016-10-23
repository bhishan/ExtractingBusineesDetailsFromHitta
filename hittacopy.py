'''
Author: Bhishan Bhandari
bbhishan@gmail.com

Description
The following script reads an excel file containing search keywords/company names. For each of these search terms,
the script makes a search on hitta.se and extracts the website link if and when available.
The excel file containing search keywords should be present in the directory where this file is present. The name 
of the file should be search_list.xlsx

Dependencies
The program uses python 2.7 version.
The program uses following external modules that needs to be installed for the program to run. 

openpyxl to read and write to excel file. In order to install it type the following command
pip install openpyxl

selenium to instantiate a browser and simulate user process in searching keywords, copying website link, etc.
pip install selenium

Chrome Driver for selenium 
https://sites.google.com/a/chromium.org/chromedriver/getting-started
Go through the above guide to have it installed or connect with the author of the script.

The following modules are by default packaged when installing python 2.7, no need to install it later. 
time	time is used to generate a pause in program for x seconds so that the browser content is fully loaded.
sys	sys is used to exit the program in case of exceptions such as file not found.
'''


import openpyxl 
import time
import sys
from selenium import webdriver 
from selenium.webdriver.common.keys import Keys

'''
Try to open the excel file containing search keywords. If file not present in the current directory, exit.
'''

try:
    wb = openpyxl.load_workbook("search_list.xlsx", read_only=False)
    ws = wb.active
except IOError:
    print "There is no excel file to read from."
    sys.exit(1)

'''
Try to instantiate a browser. If chrome driver not insalled exit.
'''

try:
    browser = webdriver.Chrome()
except:
    print "Program dependency: Please install Chrome Driver from https://sites.google.com/a/chromium.org/chromedriver/getting-started"
    sys.exit(1)

def main(base_url):
    '''
    params: base_url string
    Iterates over all the rows containing keywords in the excel sheet. If the read row is not blank, goes to 
hitta.se and writes the keyword in the search field, clicks the search icon. In case of single result, the 
result page is itself the details of the company. So the script tries to get the website link and write to excel.
In case of multiple results, the script takes the first result as a basis and goes to the link which then contains 
details of the company. Then gets website link if available and writes to excel sheet.
is  
    '''
    highest_row = ws.max_row + 1	#get the last row number from the excel sheet containing search keyword.
    for i in range(2, highest_row):	#iterate over all the keywords. Assuming keywords start from second row.
        try:
            search_keyword = ws.cell(row = i, column = 1).value #get the i th search keyword
            if search_keyword is not None:	# if the field read is not empty go inside nested block
                browser.get(base_url)	#go to hitta.se
                time.sleep(10)		#wait 10 seconds for the site to load.
                search_field = browser.find_element_by_class_name("clearable-field")	#get search field
                search_field.send_keys(search_keyword)	#write search keyword in the search input field.
                search_button = browser.find_element_by_class_name("h-icon-magnifier")	#ge the search icon element
                search_button.click()	#click the search icon to get results.
                time.sleep(20)	#wait for 20 seconds to load the search result completely.
                try:
                    '''In case of single result, the result page contains details of the company.If this was the
case, the program tries to get the home icon element which contains link to the website. Then finally writes to
 the right column of the search keyword in the same row. If this case fails, it goes to except block.'''
                    website_anchor = browser.find_element_by_class_name("h-icon-homepage")
                    website = website_anchor.get_attribute("href")
                    ws.cell(row = i, column = 2).value = website
                except:
                    try:
                        '''In case of multiple result, get the first result and click on the element to go to 
the details page. Wait 10 seconds for the page to load.'''
                        first_result_elem = browser.find_element_by_class_name("hitta-statistics-log-click")
                        first_result_elem.click()
                        time.sleep(10)
                        try:
                            '''Get the home icon which contains link to the company website. Get the link and
write to excel sheet in the corresponding row. 
                            '''
                            website_anchor = browser.find_element_by_class_name("h-icon-homepage")  
                            website = website_anchor.get_attribute("href")
                            ws.cell(row = i, column = 2).value = website
                        except:
                            print "website not available for the company."
                    except:
                        print "could not find a link"
                wb.save("search_list.xlsx")
            else:
                break	#if blank row read, break the loop.
        except:
            print"could not read excel file."

    wb.save("search_list.xlsx")
        
    
    
    

    
if __name__ == '__main__':
    main('http://www.hitta.se')		#entry point of the program

